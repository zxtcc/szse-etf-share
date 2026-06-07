# -*- coding: utf-8 -*-
"""ETF 份额历史数据的本地存储模块。

每个 ETF 的数据保存在 data/代码.xlsx，列结构（磁盘表头）：
    日期, 代码, 名称, 份额(万份), 抓取时间
程序内部统一使用规范列名「份额」（数值，单位万份）；写入 Excel 时表头显示为
「份额(万份)」并套用排版样式（加粗表头、冻结首行、千分位、列宽、筛选器），方便人工查看。
按「日期」去重（新数据覆盖旧数据），按日期升序保存。
"""

import os
import glob
import json
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 数据目录：项目根目录下的 data 文件夹（本文件位于 src/，故取上一级）
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(_PROJECT_ROOT, "data")

# 程序内部规范列名（份额单位：万份；当日变动 = 当日份额 - 前一交易日份额）
COLUMNS = ["日期", "代码", "名称", "份额", "当日变动", "抓取时间"]

# 规范列名 → Excel 磁盘表头（让单位等信息对人类可见）
HEADER_LABELS = {"份额": "份额(万份)", "当日变动": "当日变动份额(万份)"}
REVERSE_LABELS = {v: k for k, v in HEADER_LABELS.items()}

# 各列在 Excel 中的列宽
COL_WIDTHS = {
    "日期": 14, "代码": 10, "名称": 24, "份额": 18, "当日变动": 20, "抓取时间": 22,
}


def _compute_change(df_sorted):
    """根据按日期升序排列的 DataFrame，计算「当日变动份额」=当日份额-前一行份额。

    首行（无前序数据）置为 0。返回保留两位小数的 Series。
    """
    vals = pd.to_numeric(df_sorted["份额"], errors="coerce")
    return vals.diff().fillna(0).round(2)


# 汇总表的自定义排列顺序持久化文件
ORDER_FILE = os.path.join(DATA_DIR, "_order.json")


def _ensure_data_dir():
    if not os.path.isdir(DATA_DIR):
        os.makedirs(DATA_DIR, exist_ok=True)


def load_order():
    """读取汇总表自定义顺序（ETF 代码列表），文件不存在返回空列表。"""
    if os.path.isfile(ORDER_FILE):
        try:
            with open(ORDER_FILE, encoding="utf-8") as f:
                data = json.load(f)
            return [str(c) for c in data]
        except Exception:
            return []
    return []


def save_order(codes):
    """保存汇总表自定义顺序（ETF 代码列表）。"""
    _ensure_data_dir()
    with open(ORDER_FILE, "w", encoding="utf-8") as f:
        json.dump([str(c).strip() for c in codes], f, ensure_ascii=False)


def get_file_path(code):
    """返回某 ETF 代码对应的 xlsx 文件路径。"""
    return os.path.join(DATA_DIR, "%s.xlsx" % str(code).strip())


def load_df(code):
    """读取某 ETF 的历史数据为 DataFrame（规范列名），文件不存在则返回空 DataFrame。"""
    path = get_file_path(code)
    if not os.path.isfile(path):
        return pd.DataFrame(columns=COLUMNS)
    try:
        df = pd.read_excel(path, dtype={"日期": str, "代码": str})
    except Exception:
        return pd.DataFrame(columns=COLUMNS)
    # 兼容旧/新表头：把显示表头（如「份额(万份)」）还原为规范列名
    df = df.rename(columns=REVERSE_LABELS)
    # 补齐缺失列，保证列顺序统一
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[COLUMNS]


def _write_styled(df, path):
    """将 DataFrame（规范列名）写为排版美观、人类可读的 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "份额数据"

    # 表头（使用显示标签）
    headers = [HEADER_LABELS.get(c, c) for c in df.columns]
    ws.append(headers)
    # 数据行
    for _, row in df.iterrows():
        ws.append([None if pd.isna(row[c]) else row[c] for c in df.columns])

    # 表头样式：深蓝底、白字、加粗、居中
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"  # 冻结首行
    last_col = get_column_letter(len(df.columns))
    ws.auto_filter.ref = "A1:%s%d" % (last_col, ws.max_row)  # 表头筛选器

    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    for idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(col, 14)
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            if col == "份额":
                cell.number_format = "#,##0.00"  # 千分位 + 两位小数
                cell.alignment = right
            elif col == "当日变动":
                # 涨红跌绿（A 股习惯）：正数红、负数绿、零正常
                cell.number_format = "[Red]#,##0.00;[Green]-#,##0.00;0.00"
                cell.alignment = right
            elif col in ("日期", "代码"):
                cell.alignment = center

    wb.save(path)


def save_records(code, records):
    """将若干条记录写入某 ETF 的 xlsx，按日期去重后升序保存（排版美观）。

    参数：
        records: dict 列表，每条至少含 日期/代码/名称/份额
    返回：保存后该文件的总记录数。
    """
    if not records:
        return len(load_df(code))

    _ensure_data_dir()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    new_rows = []
    for rec in records:
        row = dict(rec)
        row.setdefault("代码", str(code))
        row["抓取时间"] = now
        new_rows.append(row)
    new_df = pd.DataFrame(new_rows)

    old_df = load_df(code)
    combined = pd.concat([old_df, new_df], ignore_index=True)
    # 同一日期保留最后一条（即本次新抓取的数据覆盖旧的）
    combined = combined.drop_duplicates(subset=["日期"], keep="last")
    combined = combined.sort_values("日期").reset_index(drop=True)
    # 升序排好后重算「当日变动份额」
    combined["当日变动"] = _compute_change(combined)
    combined = combined[COLUMNS]

    _write_styled(combined, get_file_path(code))
    return len(combined)


def existing_dates(code):
    """返回某 ETF 已存储的日期字符串（升序列表），用于跳过已抓取的区间。"""
    df = load_df(code)
    if df.empty:
        return []
    days = [str(x) for x in df["日期"].dropna().tolist()]
    return sorted(set(days))


def list_summary():
    """汇总本地已存储数据的所有 ETF。

    返回 dict 列表，每项含：代码, 名称, 开始日期, 结束日期, 记录数。
    排列顺序优先使用用户自定义顺序（load_order），未列入的按代码升序排在其后。
    """
    if not os.path.isdir(DATA_DIR):
        return []
    results = []
    for path in sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsx"))):
        code = os.path.splitext(os.path.basename(path))[0]
        df = load_df(code)
        df = df.dropna(subset=["日期"])
        if df.empty:
            continue
        df = df.sort_values("日期")
        dates = [str(x) for x in df["日期"].tolist()]
        names = [str(x) for x in df["名称"].dropna().tolist() if str(x).strip()]
        results.append(
            {
                "代码": code,
                "名称": names[-1] if names else "",  # 取最新一条非空名称
                "开始日期": dates[0],
                "结束日期": dates[-1],
                "记录数": len(dates),
            }
        )

    # 按用户自定义顺序排序；未列入顺序的（如新增 ETF）排在末尾并按代码升序
    order = load_order()
    pos = {c: i for i, c in enumerate(order)}
    results.sort(key=lambda r: (pos.get(r["代码"], len(order)), r["代码"]))
    return results


def delete_code(code):
    """删除某 ETF 的数据文件（data/代码.xlsx）。删除成功返回 True，文件不存在返回 False。"""
    path = get_file_path(code)
    if os.path.isfile(path):
        os.remove(path)
        return True
    return False


def load_history(code, start=None, end=None):
    """读取某 ETF 历史数据，返回按日期升序的 dict 列表，供前端图表/表格使用。

    可选 start/end（YYYY-MM-DD）按日期区间过滤。当日变动基于**完整序列**计算后再裁剪，
    使区间首行的当日变动仍是相对前一交易日的真实变化。
    """
    df = load_df(code)
    if df.empty:
        return []
    df = df.sort_values("日期").reset_index(drop=True)
    # 始终基于完整序列重算当日变动（保证旧文件与区间首行的变动都正确）
    df = df.copy()
    df["当日变动"] = _compute_change(df)
    # 按区间裁剪（日期为 YYYY-MM-DD 字符串，可直接字典序比较）
    days = df["日期"].astype(str)
    if start:
        df = df[days >= start]
    if end:
        df = df[df["日期"].astype(str) <= end]
    records = []
    for _, r in df.iterrows():
        records.append(
            {
                "日期": None if pd.isna(r["日期"]) else str(r["日期"]),
                "代码": None if pd.isna(r["代码"]) else str(r["代码"]),
                "名称": None if pd.isna(r["名称"]) else str(r["名称"]),
                "份额": None if pd.isna(r["份额"]) else float(r["份额"]),
                "当日变动": float(r["当日变动"]),
            }
        )
    return records
